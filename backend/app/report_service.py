"""Rapportages & analytics: compliance-overzicht, leveranciersscorecards,
risicosignalering en compliance-trend over tijd. Plus generieke PDF/Excel-export.

Let op: de trend wordt benaderd op basis van de huidige compliance-scores en de
aanmaakdatum van producten (er zijn geen historische snapshots). Responstijd wordt
benaderd als de gemiddelde leeftijd van afgehandelde dataverzoeken.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import List, Tuple

from sqlalchemy.orm import Session

from . import compliance, models, schemas

MAAND_NL = [
    "jan", "feb", "mrt", "apr", "mei", "jun",
    "jul", "aug", "sep", "okt", "nov", "dec",
]


# ---------- a. compliance-overzicht per wetgeving ----------
def compliance_overzicht(db: Session) -> List[schemas.ComplianceOverzichtRegel]:
    regels = []
    for wet in db.query(models.Wetgeving).order_by(models.Wetgeving.code).all():
        stats = compliance.wetgeving_stats(db, wet)
        if stats["aantal_producten"] == 0:
            continue
        cat_ids = {c.id for c in wet.categorieen}
        veld_ids = [v.id for v in wet.compliance_velden]
        ontbrekend = 0
        if cat_ids and veld_ids:
            producten = (
                db.query(models.Product)
                .filter(models.Product.categorie_id.in_(cat_ids))
                .all()
            )
            for p in producten:
                ingevuld = compliance.ingevulde_veld_ids(db, p.id)
                ontbrekend += sum(1 for vid in veld_ids if vid not in ingevuld)
        regels.append(
            schemas.ComplianceOverzichtRegel(
                code=wet.code,
                naam=wet.naam,
                aantal_producten=stats["aantal_producten"],
                compliance_percentage=stats["compliance_percentage"],
                aantal_ontbrekende_velden=ontbrekend,
            )
        )
    return regels


# ---------- b. leveranciersscorecards ----------
def _gem_responstijd(verzoeken: List[models.Dataverzoek]) -> float | None:
    afgehandeld = [
        v for v in verzoeken if v.status in ("ontvangen", "afgerond") and v.aangemaakt_op
    ]
    if not afgehandeld:
        return None
    nu = datetime.utcnow()
    dagen = [max(0, (nu - v.aangemaakt_op).days) for v in afgehandeld]
    return round(sum(dagen) / len(dagen), 1)


def scorecards(db: Session) -> List[schemas.LeverancierScorecard]:
    kaarten = []
    for lev in db.query(models.Leverancier).order_by(models.Leverancier.naam).all():
        producten = lev.producten
        if producten:
            compleetheid = round(
                sum(p.compliance_percentage or 0 for p in producten) / len(producten), 1
            )
        else:
            compleetheid = 100.0
        verzoeken = lev.dataverzoeken
        open_verzoeken = sum(
            1 for v in verzoeken if v.status in ("open", "verzonden")
        )
        kaarten.append(
            schemas.LeverancierScorecard(
                leverancier_id=lev.id,
                naam=lev.naam,
                aantal_producten=len(producten),
                compleetheid_percentage=compleetheid,
                open_verzoeken=open_verzoeken,
                gem_responstijd_dagen=_gem_responstijd(verzoeken),
            )
        )
    # slechtste compleetheid eerst (meest aandacht nodig)
    kaarten.sort(key=lambda k: (k.compleetheid_percentage, -k.open_verzoeken))
    return kaarten


# ---------- c. risicosignalering ----------
def risico(db: Session) -> List[schemas.RisicoLeverancier]:
    vandaag = date.today()
    grens = vandaag + timedelta(days=90)
    # ontbrekende velden per leverancier (gedenormaliseerde cache)
    ontbrekend_per_lev: dict[int, int] = {}
    for p in db.query(models.Product).all():
        if p.aantal_ontbrekend:
            ontbrekend_per_lev[p.leverancier_id] = (
                ontbrekend_per_lev.get(p.leverancier_id, 0) + p.aantal_ontbrekend
            )

    verzoeken = (
        db.query(models.Dataverzoek)
        .filter(
            models.Dataverzoek.deadline.isnot(None),
            models.Dataverzoek.deadline <= grens,
            models.Dataverzoek.status.in_(["open", "verzonden"]),
        )
        .order_by(models.Dataverzoek.deadline)
        .all()
    )
    # per leverancier de eerstvolgende deadline
    gezien: dict[int, schemas.RisicoLeverancier] = {}
    for v in verzoeken:
        ontbrekend = ontbrekend_per_lev.get(v.leverancier_id, 0)
        if ontbrekend == 0:
            continue
        dagen = (v.deadline - vandaag).days
        if dagen <= 30:
            cat = "30"
        elif dagen <= 60:
            cat = "60"
        else:
            cat = "90"
        if v.leverancier_id in gezien:
            continue
        gezien[v.leverancier_id] = schemas.RisicoLeverancier(
            leverancier_id=v.leverancier_id,
            naam=v.leverancier.naam if v.leverancier else "—",
            deadline=v.deadline,
            dagen_tot_deadline=dagen,
            risicocategorie=cat,
            aantal_ontbrekend=ontbrekend,
            onderwerp=v.onderwerp,
        )
    return sorted(gezien.values(), key=lambda r: (r.dagen_tot_deadline or 0))


# ---------- d. trend over tijd ----------
def _maand_label(d: date) -> str:
    return f"{MAAND_NL[d.month - 1]} {d.year}"


def trend(db: Session, maanden: int = 6) -> List[schemas.TrendPunt]:
    producten = db.query(models.Product).all()
    vandaag = date.today()
    punten: List[schemas.TrendPunt] = []
    # begin bij de eerste van de maand, `maanden` terug
    jaar, maand = vandaag.year, vandaag.month
    reeks: List[Tuple[int, int]] = []
    for _ in range(maanden):
        reeks.append((jaar, maand))
        maand -= 1
        if maand == 0:
            maand = 12
            jaar -= 1
    reeks.reverse()
    for jaar, maand in reeks:
        # einde van de maand
        if maand == 12:
            eind = date(jaar, 12, 31)
        else:
            eind = date(jaar, maand + 1, 1) - timedelta(days=1)
        bestaand = [
            p
            for p in producten
            if p.aangemaakt_op and p.aangemaakt_op.date() <= eind
        ]
        if bestaand:
            pct = round(
                sum(p.compliance_percentage or 0 for p in bestaand) / len(bestaand), 1
            )
        else:
            pct = 0.0
        punten.append(
            schemas.TrendPunt(
                maand=f"{jaar:04d}-{maand:02d}",
                label=_maand_label(eind),
                compliance_percentage=pct,
                aantal_producten=len(bestaand),
            )
        )
    return punten


def alles(db: Session) -> schemas.RapportagesData:
    return schemas.RapportagesData(
        compliance_overzicht=compliance_overzicht(db),
        scorecards=scorecards(db),
        risico=risico(db),
        trend=trend(db),
    )


# ---------- export naar tabel (headers + rijen) ----------
def _tabel_voor(db: Session, soort: str) -> Tuple[str, List[str], List[List[str]]]:
    if soort == "compliance":
        regels = compliance_overzicht(db)
        kop = ["Wetgeving", "Naam", "Producten", "Compliance %", "Ontbrekende velden"]
        rijen = [
            [r.code, r.naam, str(r.aantal_producten), f"{r.compliance_percentage}",
             str(r.aantal_ontbrekende_velden)]
            for r in regels
        ]
        return "Compliance-overzicht per wetgeving", kop, rijen
    if soort == "scorecards":
        kaarten = scorecards(db)
        kop = ["Leverancier", "Producten", "Compleetheid %", "Open verzoeken",
               "Gem. responstijd (dagen)"]
        rijen = [
            [k.naam, str(k.aantal_producten), f"{k.compleetheid_percentage}",
             str(k.open_verzoeken),
             "—" if k.gem_responstijd_dagen is None else f"{k.gem_responstijd_dagen}"]
            for k in kaarten
        ]
        return "Leveranciersscorecards", kop, rijen
    if soort == "risico":
        items = risico(db)
        kop = ["Leverancier", "Deadline", "Dagen tot deadline", "Risico (dagen)",
               "Ontbrekende velden", "Onderwerp"]
        rijen = [
            [r.naam, r.deadline.isoformat() if r.deadline else "—",
             str(r.dagen_tot_deadline), r.risicocategorie, str(r.aantal_ontbrekend),
             r.onderwerp or "—"]
            for r in items
        ]
        return "Risicosignalering", kop, rijen
    if soort == "trend":
        punten = trend(db)
        kop = ["Maand", "Compliance %", "Aantal producten"]
        rijen = [
            [p.label, f"{p.compliance_percentage}", str(p.aantal_producten)]
            for p in punten
        ]
        return "Compliance-trend", kop, rijen
    raise ValueError(f"Onbekende rapportage: {soort}")


SOORTEN = {"compliance", "scorecards", "risico", "trend"}


def bouw_xlsx(db: Session, soort: str) -> Tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    titel, kop, rijen = _tabel_voor(db, soort)
    wb = Workbook()
    ws = wb.active
    ws.title = soort[:31]
    ws.append([titel])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(kop)
    for cel in ws[3]:
        cel.font = Font(bold=True)
    for rij in rijen:
        ws.append(rij)
    for i, label in enumerate(kop, start=1):
        if i <= 26:
            ws.column_dimensions[chr(64 + i)].width = max(14, min(44, len(label) + 8))
    buf = io.BytesIO()
    wb.save(buf)
    stempel = datetime.utcnow().strftime("%Y%m%d")
    return buf.getvalue(), f"rapport-{soort}-{stempel}.xlsx"


def bouw_pdf(db: Session, soort: str) -> Tuple[bytes, str]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    titel, kop, rijen = _tabel_voor(db, soort)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=titel,
    )
    styles = getSampleStyleSheet()
    elementen = [
        Paragraph(f"PowerCompliance — {titel}", styles["Title"]),
        Paragraph(
            f"Gegenereerd op {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            styles["Normal"],
        ),
        Spacer(1, 8 * mm),
    ]
    data = [kop] + (rijen or [["Geen gegevens"] + [""] * (len(kop) - 1)])
    tabel = Table(data, repeatRows=1, hAlign="LEFT")
    tabel.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e0e0e0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elementen.append(tabel)
    doc.build(elementen)
    stempel = datetime.utcnow().strftime("%Y%m%d")
    return buf.getvalue(), f"rapport-{soort}-{stempel}.pdf"
