"""Import van leveranciers en producten uit CSV/Excel.

- Automatische kolomherkenning: headers worden genormaliseerd en gematcht aan
  bekende veldnamen (incl. veelvoorkomende synoniemen).
- Validatie: ontbrekende verplichte kolommen geven een duidelijke foutmelding.
- Compliance-analyse voor producten: extra kolommen die overeenkomen met een
  ComplianceVeld worden als waarde opgeslagen; daarna wordt per product bepaald
  of het compliant is of data mist. Koppeling aan wetgeving loopt via categorie.
"""
import csv
import io
import os
import re
from typing import List, Optional, Tuple

from sqlalchemy.orm import Session

from . import models, compliance, import_ai, audit_service


# ---------- normalisatie & mapping ----------
def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


LEVERANCIER_SYNONIEMEN = {
    "naam": {"naam", "name", "leverancier", "supplier", "bedrijf", "bedrijfsnaam", "company"},
    "contactpersoon": {"contactpersoon", "contact", "contactperson", "persoon"},
    "email": {"email", "emailadres", "mail", "emailaddress"},
    "telefoon": {"telefoon", "telefoonnummer", "tel", "phone", "phonenumber"},
    "land": {"land", "country", "landcode"},
}
LEVERANCIER_VERPLICHT = ["naam"]

PRODUCT_SYNONIEMEN = {
    "naam": {"naam", "name", "product", "productnaam", "artikelnaam", "title", "omschrijving"},
    "artikelnummer": {"artikelnummer", "artikelnr", "sku", "artikel", "articlenumber", "itemnumber", "code", "artikelcode"},
    "ean": {"ean", "barcode", "gtin", "ean13", "streepjescode"},
    "merk": {"merk", "brand", "label", "fabrikaat"},
    "beschrijving": {"beschrijving", "description", "omschrijvinglang", "productomschrijving"},
    "leverancier": {"leverancier", "supplier", "leveranciernaam", "vendor", "fabrikant"},
    "categorie": {"categorie", "category", "productgroep", "groep", "productcategorie"},
}
PRODUCT_VERPLICHT = ["naam", "leverancier"]

# Kernvelden waarnaar een kolom gemapt kan worden (sleutel → NL-label).
KERN_DOELVELDEN = [
    ("naam", "Naam"),
    ("artikelnummer", "Artikelnummer"),
    ("ean", "EAN"),
    ("merk", "Merk"),
    ("beschrijving", "Beschrijving"),
    ("leverancier", "Leverancier"),
    ("categorie", "Categorie"),
]

GELDIGE_IMPORT_MODI = {"alles", "alleen_nieuwe", "update_bestaande"}


def _bouw_synoniem_index(synoniemen: dict) -> dict:
    index = {}
    for veld, namen in synoniemen.items():
        for n in namen:
            index[_norm(n)] = veld
    return index


# ---------- bestand parsen ----------
def parse_bestand(bestandsnaam: str, inhoud: bytes) -> Tuple[List[str], List[dict]]:
    naam = (bestandsnaam or "").lower()
    if naam.endswith(".xlsx") or naam.endswith(".xlsm"):
        return _parse_xlsx(inhoud)
    return _parse_csv(inhoud)


def _parse_csv(inhoud: bytes) -> Tuple[List[str], List[dict]]:
    tekst = inhoud.decode("utf-8-sig", errors="replace")
    # delimiter detecteren (NL-Excel gebruikt vaak ;)
    monster = tekst[:2048]
    try:
        dialect = csv.Sniffer().sniff(monster, delimiters=",;\t|")
        delim = dialect.delimiter
    except csv.Error:
        delim = ";" if monster.count(";") > monster.count(",") else ","
    reader = csv.reader(io.StringIO(tekst), delimiter=delim)
    rijen = [r for r in reader if any((c or "").strip() for c in r)]
    if not rijen:
        return [], []
    headers = [h.strip() for h in rijen[0]]
    data = []
    for r in rijen[1:]:
        rij = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
        data.append(rij)
    return headers, data


def _parse_xlsx(inhoud: bytes) -> Tuple[List[str], List[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(inhoud), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        rij = {}
        for i, h in enumerate(headers):
            v = r[i] if i < len(r) else None
            rij[h] = "" if v is None else str(v).strip()
        data.append(rij)
    return headers, data


# ---------- kolommen mappen ----------
def map_kolommen(headers: List[str], synoniem_index: dict):
    """Geef terug: {header: veld} voor herkende kolommen + lijst onbekende headers."""
    herkend = {}
    onbekend = []
    for h in headers:
        veld = synoniem_index.get(_norm(h))
        if veld:
            herkend[h] = veld
        else:
            onbekend.append(h)
    return herkend, onbekend


class OntbrekendeKolommen(Exception):
    def __init__(self, kolommen):
        self.kolommen = kolommen
        super().__init__("Ontbrekende verplichte kolommen: " + ", ".join(kolommen))


# ---------- leveranciers importeren ----------
def import_leveranciers(db: Session, bestandsnaam: str, inhoud: bytes) -> dict:
    headers, rijen = parse_bestand(bestandsnaam, inhoud)
    index = _bouw_synoniem_index(LEVERANCIER_SYNONIEMEN)
    herkend, onbekend = map_kolommen(headers, index)

    ontbrekend = [v for v in LEVERANCIER_VERPLICHT if v not in herkend.values()]
    if ontbrekend:
        raise OntbrekendeKolommen(ontbrekend)

    fouten = []
    geimporteerd = 0
    for nr, rij in enumerate(rijen, start=2):  # rij 1 = header
        waarden = {veld: (rij.get(h) or "").strip() for h, veld in herkend.items()}
        if not waarden.get("naam"):
            fouten.append({"rij": nr, "bericht": "Lege of ontbrekende naam"})
            continue
        lev = models.Leverancier(
            naam=waarden.get("naam"),
            contactpersoon=waarden.get("contactpersoon") or None,
            email=waarden.get("email") or None,
            telefoon=waarden.get("telefoon") or None,
            land=waarden.get("land") or "NL",
            actief=True,
        )
        db.add(lev)
        geimporteerd += 1
    db.commit()

    return {
        "type": "leveranciers",
        "bestandsnaam": bestandsnaam,
        "aantal_rijen": len(rijen),
        "aantal_geimporteerd": geimporteerd,
        "aantal_fouten": len(fouten),
        "herkende_kolommen": herkend,
        "genegeerde_kolommen": onbekend,
        "fouten": fouten,
    }


# ---------- producten importeren + compliance-analyse ----------
def import_producten(db: Session, bestandsnaam: str, inhoud: bytes) -> dict:
    headers, rijen = parse_bestand(bestandsnaam, inhoud)
    index = _bouw_synoniem_index(PRODUCT_SYNONIEMEN)
    kern_herkend, rest = map_kolommen(headers, index)

    ontbrekend = [v for v in PRODUCT_VERPLICHT if v not in kern_herkend.values()]
    if ontbrekend:
        raise OntbrekendeKolommen(ontbrekend)

    # compliance-velden indexeren op genormaliseerde sleutel én naam
    velden = db.query(models.ComplianceVeld).all()
    veld_index = {}
    for v in velden:
        veld_index[_norm(v.sleutel)] = v
        veld_index[_norm(v.naam)] = v

    # onbekende kolommen die wél een compliance-veld zijn
    compliance_kolommen = {}  # header -> ComplianceVeld
    echt_onbekend = []
    for h in rest:
        v = veld_index.get(_norm(h))
        if v:
            compliance_kolommen[h] = v
        else:
            echt_onbekend.append(h)

    # bestaande leveranciers/categorieën op naam (case-insensitief)
    lev_op_naam = {
        l.naam.strip().lower(): l for l in db.query(models.Leverancier).all()
    }
    cat_op_naam = {
        c.naam.strip().lower(): c for c in db.query(models.Categorie).all()
    }

    herkend_overzicht = dict(kern_herkend)
    for h, v in compliance_kolommen.items():
        herkend_overzicht[h] = f"compliance: {v.wetgeving.code} · {v.naam}"

    fouten = []
    nieuwe_producten = []
    velden_ingevuld = 0

    for nr, rij in enumerate(rijen, start=2):
        kern = {veld: (rij.get(h) or "").strip() for h, veld in kern_herkend.items()}
        naam = kern.get("naam")
        lev_naam = kern.get("leverancier")
        if not naam:
            fouten.append({"rij": nr, "bericht": "Lege of ontbrekende productnaam"})
            continue
        if not lev_naam:
            fouten.append({"rij": nr, "bericht": "Leverancier ontbreekt"})
            continue
        lev = lev_op_naam.get(lev_naam.strip().lower())
        if not lev:
            fouten.append(
                {"rij": nr, "bericht": f"Leverancier '{lev_naam}' niet gevonden"}
            )
            continue

        # categorie matchen of aanmaken (bepaalt welke wetgeving van toepassing is)
        cat = None
        cat_naam = kern.get("categorie")
        if cat_naam:
            cat = cat_op_naam.get(cat_naam.strip().lower())
            if not cat:
                cat = models.Categorie(naam=cat_naam.strip())
                db.add(cat)
                db.flush()
                cat_op_naam[cat_naam.strip().lower()] = cat

        product = models.Product(
            naam=naam,
            artikelnummer=kern.get("artikelnummer") or None,
            ean=kern.get("ean") or None,
            beschrijving=kern.get("beschrijving") or None,
            leverancier_id=lev.id,
            categorie_id=cat.id if cat else None,
        )
        db.add(product)
        db.flush()

        # compliance-waarden uit extra kolommen
        for h, veld in compliance_kolommen.items():
            cel = (rij.get(h) or "").strip()
            if cel:
                db.add(
                    models.ProductComplianceWaarde(
                        product_id=product.id,
                        compliance_veld_id=veld.id,
                        waarde=cel,
                        ingevuld=True,
                    )
                )
                velden_ingevuld += 1
        nieuwe_producten.append(product)

    db.commit()

    # compliance-status per geïmporteerd product bepalen
    compliant = 0
    met_ontbrekend = 0
    for product in nieuwe_producten:
        db.refresh(product)
        stats = compliance.product_compliance(db, product)
        if stats["aantal_ontbrekend"] == 0:
            compliant += 1
        else:
            met_ontbrekend += 1

    return {
        "type": "producten",
        "bestandsnaam": bestandsnaam,
        "aantal_rijen": len(rijen),
        "aantal_geimporteerd": len(nieuwe_producten),
        "aantal_fouten": len(fouten),
        "aantal_compliant": compliant,
        "aantal_met_ontbrekende_data": met_ontbrekend,
        "aantal_velden_ingevuld": velden_ingevuld,
        "herkende_kolommen": herkend_overzicht,
        "genegeerde_kolommen": echt_onbekend,
        "fouten": fouten,
    }


# ==========================================================================
#  Slimme bulkimport (AI-kolommapping + categorisatie + preview/bevestig)
# ==========================================================================
# Hoeveel rijen we in de preview teruggeven (de tellingen zijn altijd volledig).
_PREVIEW_LIMIET = 400
# Max. aantal producten dat we in één AI-categorisatiecall meesturen.
_CATEGORIE_LIMIET = 80


def doelvelden(db: Session) -> List[dict]:
    """Alle velden waarnaar een kolom gemapt kan worden: kern + compliance."""
    velden = [
        {"sleutel": s, "label": lbl, "groep": "kern"} for s, lbl in KERN_DOELVELDEN
    ]
    veld_rijen = (
        db.query(models.ComplianceVeld)
        .join(models.Wetgeving, models.ComplianceVeld.wetgeving_id == models.Wetgeving.id)
        .order_by(models.Wetgeving.code, models.ComplianceVeld.naam)
        .all()
    )
    for cv in veld_rijen:
        code = cv.wetgeving.code if cv.wetgeving else "—"
        velden.append(
            {"sleutel": f"cf:{cv.id}", "label": f"{code} · {cv.naam}", "groep": code}
        )
    return velden


def _heuristische_mapping(db: Session, headers: List[str]) -> dict:
    """Baseline-mapping {header: veld_sleutel} via synoniemen + compliance-velden."""
    index = _bouw_synoniem_index(PRODUCT_SYNONIEMEN)
    veld_index = {}
    for v in db.query(models.ComplianceVeld).all():
        veld_index[_norm(v.sleutel)] = f"cf:{v.id}"
        veld_index[_norm(v.naam)] = f"cf:{v.id}"
    mapping = {}
    for h in headers:
        n = _norm(h)
        if n in index:
            mapping[h] = index[n]
        elif n in veld_index:
            mapping[h] = veld_index[n]
    return mapping


def bouw_mapping(db: Session, headers: List[str], rijen: List[dict]) -> dict:
    """Bepaal de kolommapping: heuristiek als basis, verrijkt met AI.

    Geeft terug: {"mapping": [{header, veld, zekerheid}], "ai_gebruikt", "ai_fout"}.
    Zorgt dat elk doelveld hoogstens één keer wordt toegewezen (eerste wint).
    """
    heur = _heuristische_mapping(db, headers)
    zekerheid = {h: (0.6 if h in heur else 0.0) for h in headers}
    per_header = {h: heur.get(h) for h in headers}

    ai_gebruikt = False
    ai_fout = None
    velden = doelvelden(db)
    ai = import_ai.ai_map_kolommen(headers, rijen[:5], velden)
    if ai:
        ai_gebruikt = True
        for h in headers:
            info = ai.get(h)
            if info and info.get("veld"):
                per_header[h] = info["veld"]
                zekerheid[h] = info.get("zekerheid", 0.7)
    elif os.environ.get("ANTHROPIC_API_KEY"):
        ai_fout = "AI-kolomanalyse mislukt; heuristische herkenning gebruikt."

    # dubbele toewijzingen ontdubbelen (hoogste zekerheid wint)
    gekozen = {}
    volgorde = sorted(headers, key=lambda h: zekerheid[h], reverse=True)
    for h in volgorde:
        veld = per_header.get(h)
        if veld and veld in gekozen:
            per_header[h] = None
        elif veld:
            gekozen[veld] = h

    mapping = [
        {"header": h, "veld": per_header.get(h), "zekerheid": round(zekerheid[h], 2)}
        for h in headers
    ]
    return {"mapping": mapping, "ai_gebruikt": ai_gebruikt, "ai_fout": ai_fout}


def _inverse(mapping: dict) -> dict:
    """{veld_sleutel: header} — eerste header die naar dat veld mapt."""
    inv = {}
    for header, veld in mapping.items():
        if veld and veld not in inv:
            inv[veld] = header
    return inv


def _bestaande_indexen(db: Session):
    """Bouw lookups van bestaande producten op EAN en artikelnummer."""
    op_ean = {}
    op_artikel = {}
    for p in db.query(models.Product).all():
        if p.ean:
            op_ean.setdefault(str(p.ean).strip().lower(), p)
        if p.artikelnummer:
            op_artikel.setdefault(str(p.artikelnummer).strip().lower(), p)
    return op_ean, op_artikel


def _match_bestaand(op_ean, op_artikel, ean, artikelnummer):
    """Geef (product, match_op) terug of (None, None). EAN gaat vóór artikelnr."""
    if ean:
        p = op_ean.get(ean.strip().lower())
        if p:
            return p, "ean"
    if artikelnummer:
        p = op_artikel.get(artikelnummer.strip().lower())
        if p:
            return p, "artikelnummer"
    return None, None


def analyseer_producten(db: Session, bestandsnaam: str, inhoud: bytes) -> dict:
    """Analyseer een geüpload productbestand zonder iets op te slaan.

    Geeft de voorgestelde kolommapping, per rij of het een nieuw product of een
    update is (match op EAN/artikelnummer) en een AI-categoriesuggestie.
    """
    headers, rijen = parse_bestand(bestandsnaam, inhoud)
    mapping_info = bouw_mapping(db, headers, rijen)
    per_header = {m["header"]: m["veld"] for m in mapping_info["mapping"]}
    inv = _inverse(per_header)

    def cel(rij, veld):
        h = inv.get(veld)
        return (rij.get(h) or "").strip() if h else ""

    op_ean, op_artikel = _bestaande_indexen(db)

    rijen_preview = []
    aantal_nieuw = aantal_update = aantal_fouten = 0
    # producten zonder categorie in het bestand → kandidaat voor AI-categorisatie
    categorie_kandidaten = []
    heeft_categorie_kolom = "categorie" in inv

    for idx, rij in enumerate(rijen):
        naam = cel(rij, "naam")
        ean = cel(rij, "ean")
        artikelnummer = cel(rij, "artikelnummer")
        leverancier = cel(rij, "leverancier")
        beschrijving = cel(rij, "beschrijving")

        if not naam:
            aantal_fouten += 1
            preview = {
                "rij_index": idx,
                "naam": None,
                "artikelnummer": artikelnummer or None,
                "ean": ean or None,
                "leverancier": leverancier or None,
                "actie": "fout",
                "melding": "Productnaam ontbreekt",
            }
            if idx < _PREVIEW_LIMIET:
                rijen_preview.append(preview)
            continue

        bestaand, match_op = _match_bestaand(op_ean, op_artikel, ean, artikelnummer)
        if bestaand:
            aantal_update += 1
            actie = "update"
        else:
            aantal_nieuw += 1
            actie = "nieuw"

        preview = {
            "rij_index": idx,
            "naam": naam,
            "artikelnummer": artikelnummer or None,
            "ean": ean or None,
            "leverancier": leverancier or None,
            "actie": actie,
            "bestaand_product_id": bestaand.id if bestaand else None,
            "match_op": match_op,
            "categorie_suggestie": cel(rij, "categorie") or None,
            "categorie_zekerheid": 1.0 if cel(rij, "categorie") else 0.0,
            "melding": None,
        }
        # alleen nieuwe producten zonder categorie krijgen een AI-suggestie
        if actie == "nieuw" and not cel(rij, "categorie"):
            categorie_kandidaten.append(
                {"index": idx, "naam": naam, "beschrijving": beschrijving}
            )
        if idx < _PREVIEW_LIMIET:
            rijen_preview.append(preview)

    # AI-categorisatie voor de kandidaten (defensief; slaat over zonder API-key)
    if categorie_kandidaten:
        bestaande_cats = [c.naam for c in db.query(models.Categorie).all()]
        suggesties = import_ai.ai_categoriseer(
            categorie_kandidaten[:_CATEGORIE_LIMIET], bestaande_cats
        )
        if suggesties:
            per_index = {p["rij_index"]: p for p in rijen_preview}
            for idx, sug in suggesties.items():
                p = per_index.get(idx)
                if p:
                    p["categorie_suggestie"] = sug["categorie"]
                    p["categorie_zekerheid"] = sug["zekerheid"]

    return {
        "bestandsnaam": bestandsnaam,
        "aantal_rijen": len(rijen),
        "headers": headers,
        "doelvelden": doelvelden(db),
        "mapping": mapping_info["mapping"],
        "rijen": rijen_preview,
        "aantal_nieuw": aantal_nieuw,
        "aantal_update": aantal_update,
        "aantal_fouten": aantal_fouten,
        "ai_gebruikt": mapping_info["ai_gebruikt"],
        "ai_fout": mapping_info["ai_fout"],
    }


def bevestig_producten(
    db: Session,
    bestandsnaam: str,
    inhoud: bytes,
    mapping: dict,
    modus: str = "alles",
    categorieen: Optional[dict] = None,
) -> dict:
    """Voer de import daadwerkelijk uit met de (gecorrigeerde) mapping.

    ``mapping``: {header: veld_sleutel}. ``modus``: alles | alleen_nieuwe |
    update_bestaande. ``categorieen``: {rij_index(str): categorienaam} — gekozen
    categorieën (uit de AI-suggestie of handmatig), koppelen automatisch de
    juiste wetgeving via de categorie.
    """
    if modus not in GELDIGE_IMPORT_MODI:
        modus = "alles"
    categorieen = categorieen or {}
    headers, rijen = parse_bestand(bestandsnaam, inhoud)
    inv = _inverse(mapping)

    def cel(rij, veld):
        h = inv.get(veld)
        return (rij.get(h) or "").strip() if h else ""

    lev_op_naam = {l.naam.strip().lower(): l for l in db.query(models.Leverancier).all()}
    cat_op_naam = {c.naam.strip().lower(): c for c in db.query(models.Categorie).all()}
    op_ean, op_artikel = _bestaande_indexen(db)

    # compliance-velden per mapping-sleutel (cf:<id>)
    compliance_kolommen = {}  # veld_sleutel -> ComplianceVeld
    for veld in set(mapping.values()):
        if veld and veld.startswith("cf:"):
            cv = db.get(models.ComplianceVeld, int(veld[3:]))
            if cv:
                compliance_kolommen[veld] = cv

    fouten = []
    aantal_nieuw = aantal_geupdatet = velden_ingevuld = aantal_gecat = 0
    geraakte_producten = []

    def zet_categorie(product, naam):
        nonlocal aantal_gecat
        if not naam:
            return
        cat = cat_op_naam.get(naam.strip().lower())
        if not cat:
            cat = models.Categorie(naam=naam.strip())
            db.add(cat)
            db.flush()
            cat_op_naam[naam.strip().lower()] = cat
        product.categorie_id = cat.id
        aantal_gecat += 1

    def zet_compliance(product, rij):
        nonlocal velden_ingevuld
        for veld_sleutel, cv in compliance_kolommen.items():
            waarde = cel(rij, veld_sleutel)
            if not waarde:
                continue
            bestaande_w = (
                db.query(models.ProductComplianceWaarde)
                .filter_by(product_id=product.id, compliance_veld_id=cv.id)
                .first()
            )
            if bestaande_w:
                bestaande_w.waarde = waarde
                bestaande_w.ingevuld = True
                bestaande_w.bron = "handmatig"
            else:
                db.add(
                    models.ProductComplianceWaarde(
                        product_id=product.id,
                        compliance_veld_id=cv.id,
                        waarde=waarde,
                        ingevuld=True,
                        bron="handmatig",
                    )
                )
            velden_ingevuld += 1

    for idx, rij in enumerate(rijen):
        naam = cel(rij, "naam")
        if not naam:
            fouten.append({"rij": idx + 2, "bericht": "Productnaam ontbreekt"})
            continue
        ean = cel(rij, "ean")
        artikelnummer = cel(rij, "artikelnummer")
        bestaand, _ = _match_bestaand(op_ean, op_artikel, ean, artikelnummer)

        # gekozen categorie: expliciete override, anders kolomwaarde
        gekozen_cat = categorieen.get(str(idx)) or categorieen.get(idx) or cel(rij, "categorie")

        if bestaand:
            if modus == "alleen_nieuwe":
                continue
            # velden bijwerken die aangeleverd zijn (niet-leeg)
            if naam:
                bestaand.naam = naam
            if artikelnummer:
                bestaand.artikelnummer = artikelnummer
            if ean:
                bestaand.ean = ean
            if cel(rij, "merk"):
                bestaand.merk = cel(rij, "merk")
            if cel(rij, "beschrijving"):
                bestaand.beschrijving = cel(rij, "beschrijving")
            if gekozen_cat:
                zet_categorie(bestaand, gekozen_cat)
            db.flush()
            zet_compliance(bestaand, rij)
            geraakte_producten.append(bestaand)
            aantal_geupdatet += 1
        else:
            if modus == "update_bestaande":
                continue
            lev_naam = cel(rij, "leverancier")
            lev = lev_op_naam.get(lev_naam.strip().lower()) if lev_naam else None
            if not lev:
                fouten.append(
                    {
                        "rij": idx + 2,
                        "bericht": (
                            f"Leverancier '{lev_naam}' niet gevonden"
                            if lev_naam
                            else "Leverancier ontbreekt (verplicht voor nieuw product)"
                        ),
                    }
                )
                continue
            product = models.Product(
                naam=naam,
                artikelnummer=artikelnummer or None,
                ean=ean or None,
                merk=cel(rij, "merk") or None,
                beschrijving=cel(rij, "beschrijving") or None,
                leverancier_id=lev.id,
            )
            db.add(product)
            db.flush()
            if gekozen_cat:
                zet_categorie(product, gekozen_cat)
            zet_compliance(product, rij)
            # nieuw product ook in de lookups zetten (voorkomt dubbele aanmaak)
            if product.ean:
                op_ean[str(product.ean).strip().lower()] = product
            if product.artikelnummer:
                op_artikel[str(product.artikelnummer).strip().lower()] = product
            geraakte_producten.append(product)
            aantal_nieuw += 1

    # audit trail: bulkimport vastleggen
    audit_service.log(
        db,
        audit_service.BULKIMPORT,
        audit_service.OBJ_IMPORT,
        object_naam=bestandsnaam,
        nieuwe_waarde=(
            f"{aantal_nieuw} nieuw, {aantal_geupdatet} bijgewerkt, "
            f"{len(fouten)} fouten"
        ),
    )
    db.commit()

    # compliance-status per geraakt product herberekenen
    for product in geraakte_producten:
        db.refresh(product)
        compliance.product_compliance(db, product)

    return {
        "type": "producten",
        "bestandsnaam": bestandsnaam,
        "aantal_rijen": len(rijen),
        "aantal_nieuw": aantal_nieuw,
        "aantal_geupdatet": aantal_geupdatet,
        "aantal_fouten": len(fouten),
        "aantal_velden_ingevuld": velden_ingevuld,
        "aantal_gecategoriseerd": aantal_gecat,
        "fouten": fouten,
    }


def template_producten_xlsx(db: Session) -> bytes:
    """Bouw een lege Excel-importtemplate met kernkolommen, wat compliance-
    kolommen, voorbeeldrijen en per-kolom instructies (als celopmerking)."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill

    # kernkolommen + tot 6 compliance-velden als voorbeeld
    kern = [
        ("Naam", "Productnaam (verplicht).", "Voorbeeldproduct A"),
        ("Artikelnummer", "Uw eigen SKU/artikelnummer (aanbevolen — voor matching).", "SKU-1001"),
        ("EAN", "EAN/GTIN-barcode (aanbevolen — voor matching).", "8712345678901"),
        ("Merk", "Merk of fabrikant.", "MerkX"),
        ("Beschrijving", "Korte productomschrijving (helpt bij categorisatie).", "Rvs schroef 4x30mm"),
        ("Leverancier", "Naam van de leverancier (verplicht, moet bestaan).", "Voorbeeld Leverancier B.V."),
        ("Categorie", "Productcategorie (optioneel — AI vult aan indien leeg).", "Bevestigingsmateriaal"),
    ]
    compliance_velden = (
        db.query(models.ComplianceVeld)
        .join(models.Wetgeving, models.ComplianceVeld.wetgeving_id == models.Wetgeving.id)
        .order_by(models.Wetgeving.code, models.ComplianceVeld.naam)
        .limit(6)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Producten"

    kop_fill = PatternFill("solid", fgColor="1a73e8")
    kop_font = Font(bold=True, color="FFFFFF")

    kolommen = list(kern)
    for cv in compliance_velden:
        code = cv.wetgeving.code if cv.wetgeving else "—"
        kolommen.append(
            (
                f"{code} · {cv.naam}",
                f"Compliance-veld voor {code}. {cv.beschrijving or ''}".strip(),
                "",
            )
        )

    for i, (label, instructie, vb1) in enumerate(kolommen, start=1):
        cel = ws.cell(row=1, column=i, value=label)
        cel.fill = kop_fill
        cel.font = kop_font
        cel.comment = Comment(instructie, "PowerCompliance")
        ws.column_dimensions[chr(64 + i) if i <= 26 else "AA"].width = max(
            14, min(40, len(label) + 6)
        )

    # twee voorbeeldrijen
    ws.append([k[2] for k in kern] + ["" for _ in compliance_velden])
    ws.append(
        ["Voorbeeldproduct B", "SKU-1002", "8712345678902", "MerkY", "Kartonnen doos 30x20", "Voorbeeld Leverancier B.V.", ""]
        + ["" for _ in compliance_velden]
    )

    # instructie-tabblad
    ws2 = wb.create_sheet("Instructies")
    ws2.append(["Kolom", "Uitleg"])
    for cel in ws2[1]:
        cel.font = Font(bold=True)
    for label, instructie, _ in kolommen:
        ws2.append([label, instructie])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
